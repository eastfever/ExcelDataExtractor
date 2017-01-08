# -*- coding: utf-8 -*-

import os
import sys
import json
import shutil
import logging
from os.path import *
from openpyxl import load_workbook

from class_schema import eDataType
from class_schema import Schema
from class_record import Record

#인코딩 설정.
reload(sys)
sys.setdefaultencoding( 'utf-8' )

# 인수 체크. 인수는 설정 파일의 경로.
if len(sys.argv) is 1 :
	print "[ExcelExport]It needs to file name to load"
	exit()

# 설정 파일에서 json 정보 읽기.
setting_file = open( sys.argv[1] )
setting_json = json.loads( setting_file.read() )
setting_file.close()

excelFolderPath = setting_json['excel_folder']
excelStringFolderPath = setting_json['excel_string_folder']

locale_codes = setting_json['locale_codes']

jsonResultFolderPath = setting_json['result_json_folder']
jsonStringResultFolderPath = setting_json['result_string_json_folder']
onlyJsonResultFolderPath = setting_json['result_only_json_folder']

csharpResultFolderPath = setting_json['result_csharp_folder']
logFolderPath = setting_json['log_folder']
extract_data_only = setting_json['extract_data_only']

# 기존 로그 파일 제거.
for root, dirs, files in os.walk( logFolderPath ):
	for fname in files :		
		logFilePath = os.path.join( root, fname )		
		if( os.path.exists( logFilePath ) ) :
			os.remove( logFilePath )

# log 폴더 새로 생성.
if not os.path.isdir( logFolderPath ) :
	os.mkdir( logFolderPath )
			
# 로깅 설정.
logFormat = "[%(asctime)-15s] (%(filename)s:%(lineno)d) %(levelname)s - %(message)s"
logging.basicConfig( filename=logFolderPath+"/log.txt", level=logging.DEBUG, format=logFormat )

logger = logging.getLogger( "ExcelLogger" )
logger.addHandler( logging.StreamHandler() )

# 엑셀로 읽은 스키마/테이블을 테이블 이름으로 매핑해서 저장하고 있는 딕셔너리.
g_recordTables = {} # 딕셔너리 형태의 데이터로 정리될 테이블들.
g_listTables = {} # 리스트 형태의 데이터로 정리될 테이블들.
g_schemaTables = {} # 데이터 스키마 정보를 담은 테이블들.
g_stringTables = {} # 국가/언어별로 분리 추출 되어야할 문자열 테이블들.
g_onlyJsonList = [] # 파싱용 C# 스크립트 없이 JSON만 파싱되면 되는 데이터 테이블들.

# 주어진 경로의 폴더 안에 있는 모든 파일 제거.
def RemoveAllFilesAtTargetFolderPath( targetPath, removeSubFolders, removeTargetExtension ) :
	for root, dirs, files in os.walk( targetPath ):
		if( removeSubFolders == True ) :
			for dir in dirs :
				joinedDirPath = os.path.join( root, dir )
				if( os.path.exists( joinedDirPath ) ) : 
					shutil.rmtree( joinedDirPath )
		for fname in files :
			targetFilePath = os.path.join( root, fname )
			if( os.path.exists( targetFilePath ) and ( removeTargetExtension == os.path.splitext( fname )[1] ) ) :
				os.remove( targetFilePath )

def GetTableNameFromSheet( excelSheet, sheetName, isStringTable ) :	
	
	# 시트명을 테이블 이름으로 사용.
	tableName = sheetName
	if( tableName is None ) :
		#logger.info( "sheet<%s> is ignored by no table name" % ( sheetName ) )
		return ""
	
	if( len( tableName ) <= 0 ) :
		#logger.info( "sheet<%s> is ignored by no table name" % ( sheetName ) )
		return ""
	
	# 테이블명이 _로 시작하지 않으면 처리 하지 않는다.
	if( tableName[0] != '_' ) :
		return ""
	
	# 테이블명에서 _제거
	tableNameLength = len( tableName )
	tableName = tableName[1:tableNameLength]
	
	# 이미 같은 이름의 테이블을 읽었는지 확인하기.
	if( g_schemaTables.get( tableName ) != None ) :
		logger.error( "table already exist - name : %s" % tableName )
		return ""
	
	# CSharp 추출없이 JSON추출만 필요한 테이블임을 확인.
	if( tableName == "OnlyJson" ) :
		return tableName
	
	if( isStringTable == True ) :
		if( g_stringTables.get( tableName ) != None ) :
			logger.error( "table already exist - name : %s" % tableName )
			return ""
	else :
		if( g_recordTables.get( tableName ) != None ) :
			logger.error( "table already exist - name : %s" % tableName )
			return ""	
	return tableName

def GetSheetSchema( schemaTable, excelSheet, fieldNameCell ) :
	# 데이터 타입이 기입된 셀의 좌표( ex. A2, B2, C2, ... ) 얻기.
	cellCoord = fieldNameCell.column + "2"
	dataTypeCell = excelSheet.cell( coordinate = cellCoord )
	
	# 첫 번째 Row의 레코드 좌표( ex. A5, B5, C5, ... ) 얻기.
	cellCoord = fieldNameCell.column + "5"
	firstRowCell = excelSheet.cell( coordinate = cellCoord )
	
	if( fieldNameCell.value == None or len( fieldNameCell.value ) == 0 ) :
		return
	
	# 스키마 정보 생성 후 보관.
	newSchema = Schema()
	newSchema.columnName = fieldNameCell.column
	newSchema.fieldName = fieldNameCell.value
	newSchema.SetDataType( dataTypeCell.value, firstRowCell.value )
	
	# ID필드 체크.
	cellCoord = fieldNameCell.column + "4"
	uniqueTypeCell = excelSheet.cell( coordinate = cellCoord )		
	if uniqueTypeCell.value == 1 :
		newSchema.SetAsUniqueField()
	schemaTable[ fieldNameCell.value ] = newSchema
	#logger.info( "%s data type - %s" % ( dataTypeCell.value, newSchema.dataType ) )
	
# 엑셀의 각 시트에서 데이터를 읽어, 테이블 형태로 정리한다.
def LoadTablesFromSheets( loadedExcel, excelFileName, sheetName, isStringTable ) :
		
	excelSheet = loadedExcel[sheetName]
	
	# 테이블 이름 얻기. 이름을 확인할 수 없으면 더 이상 진행 안 함.
	tableName = GetTableNameFromSheet( excelSheet, sheetName, isStringTable )
	
	# 시트명을 테이블 이름으로 사용하기. 이름을 확인할 수 없으면 더 이상 진행 안 함.
	#tableName = sheetName
	if( len( tableName ) <= 0 ) :
		return
	
	# 시트 이름이 data일 경우 테이블 생성 안하고 건너 뛴다.
	if( tableName == 'data' ) :
		return
	
	#logger.info( "sheet<%s> processing start" % ( sheetName ) )
	
	# 테이블 스키마 분석.
	schemaTable = {}
	fieldNameCells = excelSheet.rows[2]
	for fieldNameCell in fieldNameCells :		
		# 비어있는 셀은 무시.
		if( fieldNameCell.value is None ) :
			continue		
		GetSheetSchema( schemaTable, excelSheet, fieldNameCell )
	
	# Row단위로 레코드 생성해서 Row번호를 ID로 갖는 딕셔너리에 담아 둠.
	recordTable = {}
	rowCount = len( excelSheet.rows )
	for rowNo in range( 5, rowCount + 1 ) :			
		newRecord = Record()
		for schema in schemaTable.values() :
			cellCoord = "%s%d" % ( schema.columnName, rowNo )
			selectedCell = excelSheet.cell( coordinate = cellCoord )
			if( selectedCell.value != None ) :
				newRecord.AddField( schema, selectedCell.value )
				
		if( len( newRecord.fieldTable ) > 0 ) :
			recordTable[ rowNo ] = newRecord
	
	logger.info( "table loading completed - %s" % ( tableName ) )
	
	# Json추출만 필요한 시트에 대해서는 별도 처리를 위해 추출에 필요한 정보 따로 모아 두기
	if( tableName == "OnlyJson" ) :
		OnlyJsonInfo = {}
		OnlyJsonInfo[ "file_name" ] = "%s_%s" % ( excelFileName, sheetName )
		OnlyJsonInfo[ "record_table" ] = recordTable
		g_onlyJsonList.append( OnlyJsonInfo )
		return
	
	# 모든 곳에서 접근 할 수 있는 테이블 딕셔너리에 이번 씨트에서 읽은 테이블 저장
	if( isStringTable == True ) :
		g_stringTables[ tableName ] = recordTable
	else :	
		idSchema = GetUniqueSchemaWithTableName( schemaTable )
		if( None == idSchema ) :
			# UniqueID가 설정되지 않은 엑셀은 리스트 형태로 데이터가 추출 될 수 있도록 한다.
			g_listTables[ tableName ] = recordTable
		else :
			g_recordTables[ tableName ] = recordTable
	g_schemaTables[ tableName ] = schemaTable

def CreateJson( targetPath, recordTable ) :
	#JSON으로 변환하기 쉽게 List/Dictionary형태로만 이루어진 데이터를 만든다.[ {}, {}, {} ] 
	listForJSON = []		
	for record in recordTable.values() :
		newDictionary = {}
		for fieldName in record.fieldTable.keys() :
			newDictionary[ fieldName ] = record.fieldTable[ fieldName ].value			
		listForJSON.append( newDictionary )	

	recordJSON = json.dumps( listForJSON, indent=4, sort_keys=True, ensure_ascii=False )	
	resultFile = open( targetPath, 'w' )
	resultFile.write( recordJSON.decode( 'utf-8' ) )
	resultFile.close()
	logger.info( "json created - %s" % ( targetPath ) )

def CreateJsonOnly( targetPath ) :
	for extractInfo in g_onlyJsonList :
		record_table = extractInfo[ "record_table" ]		
		resultPath = os.path.join( targetPath, ( "%s.txt" % extractInfo[ "file_name" ] ) )
		CreateJson( resultPath, record_table )

# 테이블 단위로 데이터 json파일을 생성한 후,
# 각 테이블의 데이터 형식 정보를 담은 스키마 json파일을 생성한다.
def CreateJsonByTables( targetPath ) :

	# 테이블 단위로 데이터 json파일 생성.
	for tableName in g_recordTables.keys() :
		table = g_recordTables[ tableName ]
		resultPath = os.path.join( targetPath, ( "%s.txt" % tableName ) )
		CreateJson( resultPath, table )
		
	for tableName in g_listTables.keys() :
		table = g_listTables[ tableName ]
		resultPath = os.path.join( targetPath, ( "%s.txt" % tableName ) )
		CreateJson( resultPath, table )
	# 각 테이블의 스키마 정보를 담은 json파일 생성.
	#dictForJSON = {}
	#for tableName in g_schemaTables.keys() :
		#listForSchema = [];
		#schemaTable = g_schemaTables[ tableName ]
		#for schema in schemaTable.values() :		
			#newDictionary = {}
			#newDictionary[ "fieldName" ] = schema.fieldName
			#newDictionary[ "dataType" ] = schema.dataType
			#listForSchema.append( newDictionary )
		#dictForJSON[ tableName ] = listForSchema
	
	#schemaJSON = json.dumps( dictForJSON, indent=4, sort_keys=True )	
	#resultFile = open( os.path.join( targetPath, "table_schema.txt" ), "w" )
	#resultFile.write( schemaJSON )
	#resultFile.close()
	#logger.info( "table schema completed" )

# 각 언어 코드마다 문자열 json파일 생성 후, 언어 코드별로 분류된 폴더에 위치 시킨다.
def CreateLocalizedStringJson( targetPath ) :
	
	# 언어 코드 단위 별로 처리.
	for localeCode in locale_codes : 		
		
		#코드명과 일치하는 서브폴더가 있는지 확인하고 없으면 만든다.
		localeFolder = os.path.join( targetPath, localeCode )
		if not os.path.isdir( localeFolder ) :
			os.mkdir( localeFolder )
		
		# 테이블 단위로 데이터 json파일 생성.
		for tableName in g_stringTables.keys() :
			table = g_stringTables[ tableName ]
			schema = g_schemaTables[ tableName ]
			#JSON으로 변환하기 쉽게 List/Dictionary형태로만 이루어진 데이터를 만든다.[ {}, {}, {} ] 
			listForJSON = []		
			for record in table.values() :			
				newDictionary = {}
				for fieldName in record.fieldTable.keys() :					
					# ID와 현재 처리중인 국가 코드와 매칭되는 필드만 찾아서 데이터를 만든다.
					fieldSchema = schema[ fieldName ]
					if( ( fieldName != localeCode ) and ( True != fieldSchema.isUnique ) ) :
						continue
					
					# JSON의 field를 id와 string으로 고정.
					newFieldName = "string"
					if( True == fieldSchema.isUnique ) :
						newFieldName = "id"
					newDictionary[ newFieldName ] = record.fieldTable[ fieldName ].value
				
				listForJSON.append( newDictionary )	
			
			recordJSON = json.dumps( listForJSON, indent=4, sort_keys=True, ensure_ascii=False )
			
			resultPath = os.path.join( localeFolder, ( "%s.txt" % tableName ) )		
			resultFile = open( resultPath, 'w' )
			resultFile.write( recordJSON.decode( 'utf-8' ) )
			resultFile.close()
			logger.info( "string json completed - %s" % ( resultPath ) )
		logger.info( "localized string json completed - %s" % ( localeCode ) )
		logger.info( "" );

# ExcelTableLoader의 Getter메서드 작성
def AppendGetterMethodOfExcelTableLoader( codeLineList, tableName ) :
	codeLineList.append( "\n" )
	codeLineList.append( "\tpublic %s Get%s()\n" % ( tableName, tableName ) )
	codeLineList.append( "\t{\n" )
	
	codeLineList.append( "\t\tif( !_tables.ContainsKey( \"%s\" ) )\n" % tableName )
	codeLineList.append( "\t\t{\n" )
	codeLineList.append( "\t\t\t%s newTable = new %s();\n" % ( tableName, tableName ) )
	codeLineList.append( "\t\t\tnewTable.Load( \"%s\" );\n" % tableName )
	codeLineList.append( "\t\t\t_tables.Add( \"%s\", newTable );\n" % tableName )
	codeLineList.append( "\t\t}\n" )
	
	codeLineList.append( "\t\tExcelTableBase tableBase = _tables[ \"%s\" ];\n" % tableName )
	codeLineList.append( "\t\treturn tableBase as %s;\n" % tableName )
	
	codeLineList.append( "\t}\n" )
	
# DataTableLoader.cs 제작.
def CreateExcelTableLoaderDotCS( targetPath ) :	
	codeLineList = []
	
	#using
	codeLineList.append( "using UnityEngine;\n" )
	codeLineList.append( "using System.Collections;\n" )
	codeLineList.append( "using System.Collections.Generic;\n" )
	codeLineList.append( "\n" )
	
	#class 선언.
	codeLineList.append( "public class ExcelTableLoader\n" )
	codeLineList.append( "{\n" )
	
	#싱글턴.
	#codeLineList.append( "\tstatic private ExcelTableLoader s_instance = null;\n" )
	#codeLineList.append( "\tpublic static ExcelTableLoader Instance\n" )
	#codeLineList.append( "\t{\n" )
	#codeLineList.append( "\t\tget\n" )
	#codeLineList.append( "\t\t{\n" )
	#codeLineList.append( "\t\t\tif( null == s_instance )\n" )
	#codeLineList.append( "\t\t\t{\n" )
	#codeLineList.append( "\t\t\t\ts_instance = new ExcelTableLoader();\n" )	
	#codeLineList.append( "\t\t\t}\n" )
	#codeLineList.append( "\t\t\treturn s_instance;\n" )
	#codeLineList.append( "\t\t}\n" )
	#codeLineList.append( "\t}\n" )
	#codeLineList.append( "\n" )
	#codeLineList.append( "\tprivate ExcelTableLoader(){}" )
	#codeLineList.append( "\n" )
	
	#테이블 딕셔너리.
	codeLineList.append( "\n" )
	codeLineList.append( "\tprivate Dictionary<string, ExcelTableBase> _tables = new Dictionary<string, ExcelTableBase>();" )
	codeLineList.append( "\n" )
	
	#클리어 처리.
	codeLineList.append( "\n" )
	codeLineList.append( "\tpublic void Clear() { _tables.Clear(); }" )
	codeLineList.append( "\n" )
	
	#엑셀에서 읽은 모든 테이블의 Get메서드 추가.
	for tableName in g_recordTables.keys() :
		AppendGetterMethodOfExcelTableLoader( codeLineList, tableName )
	for tableName in g_listTables.keys() :
		AppendGetterMethodOfExcelTableLoader( codeLineList, tableName )
	
	# class 선언 종료.
	codeLineList.append( "}\n" )	
	
	# .cs파일 생성.
	resultCS = "".join( codeLineList )	
	resultFile = open( os.path.join( targetPath, "ExcelTableLoader.cs" ), "w" )
	resultFile.write( resultCS )
	resultFile.close()
	logger.info( "ExcelTableLoader.cs completed" )

def recordMappingAsDataType( dataType, fieldName ) :
	if( dataType == eDataType.Int ) :		
		return ( "int.Parse( record[ \"%s\" ].ToString() )" % fieldName )
	elif( dataType == eDataType.Float ) :
		return ( "float.Parse( record[ \"%s\" ].ToString() )" % fieldName )
	elif( dataType == eDataType.String ) :	
		return ( "record[ \"%s\" ] as string" % fieldName )	
	elif( dataType == eDataType.Table ) :
		return ( "DataTableLoader.Instance.Get%s()" % fieldName )
	elif( dataType == eDataType.IntList ) :
		return "new List<int>()"
	elif( dataType == eDataType.StringList ) :
		return "new List<string>()"

def makeList( schemaAsParam, codeLineListAsParam ) :
	codeLineListAsParam.append( "\t\t\tList<object> stringArray = MiniJSON.Json.Deserialize( record[ \"%s\" ].ToString() ) as List<object>;\n" % schemaAsParam.fieldName )
	codeLineListAsParam.append( "\t\t\tforeach( object listObject in stringArray )\n" )
	codeLineListAsParam.append( "\t\t\t{\n" );
	if( schemaAsParam.dataType == eDataType.IntList ) :
		codeLineListAsParam.append( "\t\t\t\tnewData.%s.Add( int.Parse( listObject.ToString() ) );\n" % schemaAsParam.fieldName );
	elif( schemaAsParam.dataType == eDataType.StringList ) :	
		codeLineListAsParam.append( "\t\t\t\tnewData.%s.Add( listObject.ToString() );\n" % schemaAsParam.fieldName );
	codeLineListAsParam.append( "\t\t\t}\n" );
	
#def recordMappingAsDataType( dataType, fieldName ) :
#	if( dataType == eDataType.Int ) :		
#		return ( "( int )record[ \"%s\" ]" % fieldName )
#	elif( dataType == eDataType.Float ) :
#		return ( "( float )record[ \"%s\" ]" % fieldName )
#	elif( dataType == eDataType.String ) :	
#		return ( "record[ \"%s\" ] as string" % fieldName )
#	elif( dataType == eDataType.Table ) :
#		return ( "DataTableLoader.Instance.Get%s()" % fieldName )

def GetUniqueSchemaWithTableName( passedSchemaTable ):
	# Unique타입 스키마 찾기
	for key, value in passedSchemaTable.iteritems():
		if value.isUnique is True:
			return value			
	return None

# CSharp 코드의 상단 공통 부분 작성
def AppendHeaderOfCSharp( codeLineList, tableName ) :
	#using
	codeLineList.append( "using UnityEngine;\n" )
	codeLineList.append( "using System.Collections;\n" )
	codeLineList.append( "using System.Collections.Generic;\n" )
	codeLineList.append( "\n" )
	
	# 레코드 class 선언.
	recordClassName = tableName + "Record"
	codeLineList.append( "public class %s\n" % recordClassName )
	
	# 각 필드의 자료형과 필드명 정의.
	codeLineList.append( "{\n" )	
	
	for schema in g_schemaTables[ tableName ].values() :
		if( schema.dataType != eDataType.Table ) :
			codeLineList.append( "\tpublic %s %s;\n" % ( schema.TypeEnumString(), schema.fieldName ) )
		else :
			codeLineList.append( "\tpublic %s %sLink;\n" % ( schema.fieldName, schema.fieldName ) )
	
	# 레코드 class 선언 종료.
	codeLineList.append( "}\n" )
	
	return recordClassName

# CSharp 코드의 RecordFromDictionary메서드 구현부 추가
def AppendRecordFromDictionaryMethodOfCSharp( codeLineList, tableName ) :
	
	recordClassName = tableName + "Record"	
	codeLineList.append( "\tprotected override %s RecordFromDictionary( Dictionary<string, object> record )\n" % recordClassName )
	codeLineList.append( "\t{\n" )
			
	codeLineList.append( "\t\t%s newData = new %s();\n" % ( recordClassName, recordClassName ) )	
	for schema in g_schemaTables[ tableName ].values() :
		codeLineList.append( "\t\tif( record.ContainsKey( \"%s\" ) )\n" % schema.fieldName )
		codeLineList.append( "\t\t{\n" )
		assignCode = "\t\t\tnewData.%s = %s;\n"
		if( schema.dataType == eDataType.Table ) :
			assignCode = "\t\t\tnewData.%sLink = %s;\n"
		codeLineList.append( assignCode % ( schema.fieldName, recordMappingAsDataType( schema.dataType, schema.fieldName ) ) )
		
		if( schema.dataType == eDataType.IntList or schema.dataType == eDataType.StringList ) :			
			makeList( schema, codeLineList )
		codeLineList.append( "\t\t}\n" )
	
	codeLineList.append( "\t\treturn newData;\n" )
	codeLineList.append( "\t}\n" )
	
# 각 테이블의 자료구조용 C#파일 생성.
def CreateTableDataStructures( targetPath ) :
	
	for tableName in g_recordTables.keys() :
	
		# Unique타입 스키마 찾기
		idSchema = GetUniqueSchemaWithTableName( g_schemaTables[ tableName ] )
		if( None == idSchema ) :
			logger.error( "not found unique data field in %s table" % tableName )
			continue
		
		if( idSchema.dataType != eDataType.Int and idSchema.dataType != eDataType.String  ) :
			logger.error( "ID field data type must be int or string in %s table" % tableName )
			continue
		
		codeLineList = []
		recordClassName = AppendHeaderOfCSharp( codeLineList, tableName )
		codeLineList.append( "\n" )
		
		# 테이블 class 선언.		
		codeLineList.append( "public class %s : ExcelTable<%s, %sRecord>\n" % ( tableName, idSchema.TypeEnumString(), tableName ) )
		codeLineList.append( "{\n" )
		
		# KeyFromDictionary 메서드. ####################################
		codeLineList.append( "\tprotected override %s KeyFromDictionary( Dictionary<string, object> record )\n" % idSchema.TypeEnumString() )
		codeLineList.append( "\t{\n" )
		
		# if( !record.ContainsKey( "ID" ) ).
		codeLineList.append( "\t\tif( !record.ContainsKey( \"%s\" ) )\n" % idSchema.fieldName )
		codeLineList.append( "\t\t{\n" )				
		if( idSchema.dataType == eDataType.Int ) :
			codeLineList.append( "\t\t\tErrorLog( 0, \"ID field not found\" );\n" )
			codeLineList.append( "\t\t\treturn -1;\n" )		
		elif( idSchema.dataType == eDataType.String ) :
			codeLineList.append( "\t\t\tErrorLog( \"\", \"ID field not found\" );\n" )
			codeLineList.append( "\t\t\treturn \"\";\n" )
		codeLineList.append( "\t\t}\n" )
		
		# return.
		codeLineList.append( "\t\treturn %s;\n" % recordMappingAsDataType( idSchema.dataType, idSchema.fieldName ) )
		
		# KeyFromDictionary 메서드 종료. ####################################
		codeLineList.append( "\t}\n" )
		codeLineList.append( "\n" )		
		
		# RecordFromDictionary 메서드 
		AppendRecordFromDictionaryMethodOfCSharp( codeLineList, tableName )
		
		# 테이블 class 선언 종료.
		codeLineList.append( "}\n" )
		
		# .cs파일 생성.
		resultCS = "".join( codeLineList )				
		resultFile = open( os.path.join( targetPath, ( "%s.cs" % tableName ) ), "w" )
		resultFile.write( resultCS )
		resultFile.close()
		logger.info( "table csharp completed - %s.cs" % ( tableName ) )

# 리스트 형태의 데이터를 다룰 C# 코드 생성.
def CreateListDataStructures( targetPath ) :
	for tableName in g_listTables.keys() :
	
		codeLineList = []
		recordClassName = AppendHeaderOfCSharp( codeLineList, tableName )
		codeLineList.append( "\n" )
		
		# 테이블 class 선언.		
		codeLineList.append( "public class %s : ExcelRecordList<%sRecord>\n" % ( tableName, tableName ) )
		codeLineList.append( "{\n" )
		
		# RecordFromDictionary 메서드 
		AppendRecordFromDictionaryMethodOfCSharp( codeLineList, tableName )
		
		# 테이블 class 선언 종료.
		codeLineList.append( "}\n" )
		
		# .cs파일 생성.
		resultCS = "".join( codeLineList )				
		resultFile = open( os.path.join( targetPath, ( "%s.cs" % tableName ) ), "w" )
		resultFile.write( resultCS )
		resultFile.close()
		logger.info( "list csharp completed - %s.cs" % ( tableName ) )
		
# 엑셀 파일인지 확인
def IsExcelFile( fileName ) :
	if ".xlsx" not in fileName:
		return False
	if ".meta" in fileName:
		return False
	if ".svn-base" in fileName:
		return False
	return True

# 엑셀에서 데이터 읽기
def LoadDataFromExcels( rootPath, fileName, isStringTable ) :	
	if ( IsExcelFile( fileName ) == False ) :
		return
	excelFilePath = os.path.join( rootPath, fileName )
	logger.info( "extracting data from - %s" % excelFilePath )
	loadedExcel = load_workbook( excelFilePath, data_only=True )

	# 각 엑셀 시트마다 돌면서 테이블 단위로 데이터 로딩.
	for sheet in loadedExcel.get_sheet_names() :
		LoadTablesFromSheets( loadedExcel, os.path.splitext( fileName )[0], sheet, isStringTable )
	logger.info( "" )
	
# 엑셀 데이터 처리 시작. ################################################################
logger.info( "" )
logger.info( "--- START ---" )
logger.info( "" )

# 엑셀 폴더 안의 존재하는 모든 엑셀 파일의 데이터를 읽는다.
for root, dirs, files in os.walk( excelFolderPath ):	
	for fname in files:		
		LoadDataFromExcels( root, fname, False )

# 문자열 엑셀 폴더 안의 존재하는 모든 엑셀 파일의 데이터를 읽는다.
for root, dirs, files in os.walk( excelStringFolderPath ):
	for fname in files:		
		LoadDataFromExcels( root, fname, True )
				
# JSON추출만 필요한 엑셀로 부터 JSON파일 생성.
RemoveAllFilesAtTargetFolderPath( onlyJsonResultFolderPath, False, ".txt" )
CreateJsonOnly( onlyJsonResultFolderPath )
				
# 각 테이블 단위로 JSON파일 생성.
RemoveAllFilesAtTargetFolderPath( jsonResultFolderPath, False, ".txt" )
CreateJsonByTables( jsonResultFolderPath )

# 각 문자열 테이블 단위로 JSON파일 생성.
RemoveAllFilesAtTargetFolderPath( jsonStringResultFolderPath, False, ".txt" )
CreateLocalizedStringJson( jsonStringResultFolderPath )

if ( extract_data_only != 1 ) :
	# ExcelTableLoader.cs
	RemoveAllFilesAtTargetFolderPath( csharpResultFolderPath, False, ".cs" )
	CreateExcelTableLoaderDotCS( csharpResultFolderPath )

	# 딕셔너리 형태의 자료구조용 C#파일 생성.
	CreateTableDataStructures( csharpResultFolderPath )

	# 리스트 형태의 자료구조용 C#파일 생성.
	CreateListDataStructures( csharpResultFolderPath )

#참고 : DataTable.cs는 파이썬이 생성하지 않고, 직접 작성했다.
logger.info( "" )
logger.info( "--- END ---" )
